import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from datetime import timedelta, datetime
import export_pdf
import urllib.parse, os
from openpyxl import load_workbook
from openpyxl.styles import Alignment


st.set_page_config(page_title='Prisma Cloud Report Dashboard', page_icon=':bar_chart:', layout='wide')
if st.button("Clear Cache"):
    st.cache_data.clear()

if "filename" in st.query_params:
    filename = st.query_params["filename"]
    
    if os.path.exists(filename):
        # @st.cache_data
        def load_data():

            df = pd.read_excel(
                io=filename,
                engine='openpyxl',
            )
            df.index += 1
            return df

        def process_data(df):
            # Unique values for filtering
            attack_type_options = df["AttackType"].unique().tolist()
            namespace_options = df["Namespace"].unique().tolist()
            cluster_options = df["Cluster"].unique().tolist()
            container_options = df["containerName"].unique().tolist()

            # 'Select All' option
            all_option = "Select All"
            attack_type_choices = [all_option] + attack_type_options
            namespace_choices = [all_option] + namespace_options
            cluster_choices = [all_option] + cluster_options
            container_choices = [all_option] + container_options

            # Sidebar filter for AttackType
            attack_type = st.sidebar.multiselect("AttackType:", options=attack_type_choices, default=[], key="attack_type_filter")

            # Apply 'Select All' logic for AttackType
            attack_type_condition = attack_type_options if all_option in attack_type or not attack_type else attack_type

            # Step 1: Filter by AttackType to dynamically update Namespace options
            if attack_type:
                filtered_namespaces = df[df["AttackType"].isin(attack_type_condition)]["Namespace"].unique().tolist()
            else:
                filtered_namespaces = namespace_options

            # Update Namespace options
            namespace_choices = [all_option] + filtered_namespaces
            namespace = st.sidebar.multiselect("Namespace:", options=namespace_choices, default=[], key="namespace_filter")

            # Apply 'Select All' logic for Namespace
            namespace_condition = namespace_options if all_option in namespace or not namespace else namespace

            # Step 2: Filter by both AttackType and Namespace to dynamically update Cluster options
            filtered_df = df[
                (df["AttackType"].isin(attack_type_condition)) & 
                (df["Namespace"].isin(namespace_condition))
            ]
            filtered_clusters = filtered_df["Cluster"].unique().tolist()

            # Update Cluster options
            cluster_choices = [all_option] + filtered_clusters
            cluster = st.sidebar.multiselect("Cluster:", options=cluster_choices, default=[], key="cluster_filter")

            # Apply 'Select All' logic for Cluster
            cluster_condition = cluster_options if all_option in cluster or not cluster else cluster

            filtered_df = filtered_df[(filtered_df["Cluster"].isin(cluster_condition))]
            filtered_containers = filtered_df["containerName"].unique().tolist()

            # Update Container options
            container_choices = [all_option] + filtered_containers
            container = st.sidebar.multiselect("Container:", options=container_choices, default=[], key="container_filter")

            # Apply 'Select All' logic for Container
            container_condition = container_options if all_option in container or not container else container

            # Final selection: Filter the dataframe based on the selected conditions
            df_selection = df.query(
                 "AttackType in @attack_type_condition and Namespace in @namespace_condition and Cluster in @cluster_condition and containerName in @container_condition"
            )
            filters = {
                "Attack Type": attack_type if attack_type else "All",
                "Namespace": namespace if namespace else "All",
                "Cluster": cluster if cluster else "All",
            }
            if df_selection.empty:
                st.warning("No data available for the selected filters. Please adjust your filters.")
                return None
            cluster_attack_counts = (
            df_selection.groupby("Cluster")["AttackType"]
                .nunique()
                .sort_values(ascending=False)
                .head(5)
                .reset_index(name="UniqueAttackTypes")
                .sort_values(by="UniqueAttackTypes", ascending=True)
            )
            container_unique_attack_counts = (
                df_selection.groupby("containerName")["AttackType"]
                .nunique()
                .sort_values(ascending=False)
                .head(5)
                .reset_index()
                .rename(columns={"AttackType": "UniqueAttackTypes", "containerName": "Container"})
                .sort_values(by="UniqueAttackTypes", ascending=True)
            )
            container_attack_counts = (
                df_selection.groupby("containerName")["AttackType"]
                .count()
                .sort_values(ascending=False)
                .head(5)
                .reset_index()
                .rename(columns={"AttackType": "TotalAttacks", "containerName": "Container"})
                .sort_values(by="TotalAttacks", ascending=True)
            )
            attack_type_counts = (
                df_selection["AttackType"]
                .value_counts()
                .head(5)
                .reset_index()
                .rename(columns={"index": "AttackType", "count": "TotalOccurrences"})
            )
        

            return df_selection, cluster_attack_counts, container_unique_attack_counts, container_attack_counts, attack_type_counts, filters

        @st.cache_data
        def create_summary_dataframe(df):
            df_summary = df.groupby(['containerName', 'imageName', 'Namespace'], as_index=False).agg({
                'Message': lambda x: '\n\n'.join(x.unique())
            })
            df_summary.rename(columns={
                'containerName': 'Container',
                'imageName': 'Image',
                'Namespace': 'Namespace',
                'Message': 'Message'
            }, inplace=True)
            
            return df_summary
        @st.cache_data
        def plot_chart_cluster_most_unique_attack(cluster_unique_attack):
            if cluster_unique_attack is not None and not cluster_unique_attack.empty:
                fig_url = px.bar(
                    cluster_unique_attack,
                    x="UniqueAttackTypes",
                    y="Cluster",
                    labels={"UniqueAttackTypes": "Number of Unique Attack Types", "Cluster": "Cluster"},
                    title="<b>Top Clusters with the Most Unique Attack Types</b>",
                    orientation="h",
                    template="plotly_white",
                    text="UniqueAttackTypes",
                )
                fig_url.update_traces(textposition="outside")
                fig_url.update_layout(title_x=0, xaxis_title="Number of Unique Attack Types", yaxis_title="Cluster")
                return fig_url
            else:
                return None
        @st.cache_data
        def plot_chart_container_most_attacks(container_attack_counts):
            if df_selection is not None and not df_selection.empty:
                fig = px.bar(
                    container_attack_counts,
                    x="TotalAttacks",
                    y="Container",
                    labels={"TotalAttacks": "Number of Attacks", "Container": "Container Name"},
                    title="<b>Top Containers with the Most Attacks</b>",
                    orientation="h",
                    template="plotly_white",
                    text="TotalAttacks"
                )
                fig.update_traces(textposition="outside")
                fig.update_layout(
                    title={"text": "<b>Top Containers with the Most Attacks</b>", "x": 0, "xanchor": "left"},
                    xaxis_title="Number of Attacks",
                    yaxis_title="Container Name"
                )
                return fig
            else:
                return None
        @st.cache_data
        def plot_chart_container_most_unique_attack(container_unique_attack_counts):
            if container_unique_attack_counts is not None:
                fig = px.bar(
                    container_unique_attack_counts,
                    x="UniqueAttackTypes",
                    y="Container",
                    labels={"UniqueAttackTypes": "Number of Unique Attack Types", "Container": "Container Name"},
                    title="<b>Top Containers with the Most Unique Attack Types</b>",
                    orientation="h",
                    template="plotly_white",
                    text="UniqueAttackTypes"
                )
                fig.update_traces(textposition="outside")
                fig.update_layout(title_x=0, xaxis_title="Number of Unique Attack Types", yaxis_title="Container Name")
                return fig
            return None
        
        @st.cache_data
        def plot_top_attack_types(attack_type_counts):
            if attack_type_counts is not None:
                attack_type_counts = attack_type_counts.sort_values(by="TotalOccurrences", ascending=True)                
                fig = px.bar(
                    attack_type_counts,
                    x="TotalOccurrences",
                    y="AttackType",
                    labels={"TotalOccurrences": "Number of Occurrences", "AttackType": "Attack Type"},
                    title="<b>Top 5 Attack Types</b>",
                    orientation="h",
                    template="plotly_white",
                    text="TotalOccurrences"
                )
                fig.update_traces(textposition="outside")
                fig.update_layout(
                    title={"text": "<b>Top 5 Attack Types</b>", "x": 0, "xanchor": "left"},
                    xaxis_title="Number of Occurrences",
                    yaxis_title="Attack Type"
                )
                return fig
            else:
                return None
        
        def generate_message_content(df_summary):
            message_content = []
            for _, row in df_summary.iterrows():
                container = f"**Container:** {row['Container']}"
                image = f"**Image:** {row['Image']}"
                namespace = f"**Namespace:** {row['Namespace']}"
                message = f"**Message:**\n{row['Message']}"  # Add line breaks for formatting
                
                # Collect content as a list of strings
                message_content.append(f"{container}\n{image}\n{namespace}\n{message}\n")
                return "\n\n".join(message_content)
            
        def generate_excel_cluster(df):
            try:
                filename = "test_cluster_output.xlsx"
                headers = {
                    "containerName": "Container",
                    "imageName": "Image",
                    "Namespace": "Namespace",
                    "Message": "Message",
                }
                grouped = df.groupby(['Cluster', 'containerName', 'imageName', 'Namespace'])['Message'].apply(
                    lambda x: '\n'.join(x.unique())
                ).reset_index()
                with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                    for cluster, group in grouped.groupby("Cluster"):
                        group.drop(columns='Cluster', inplace=True)
                        group.to_excel(writer, sheet_name=cluster, index=False)
                
                wb = load_workbook(filename)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                            
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column].width = adjusted_width
                        
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, max_row=ws.max_row):
                        for cell in row:
                            if cell.column == 4:
                                cell.alignment = Alignment(wrap_text=True)
                wb.save(filename)
                return filename
            except Exception as e:
                print(f"Error generating Excel cluster: {e}")
                return None
            
        # --- SIDEBAR ---
        st.sidebar.header("Filter")
        df = load_data()
        df_selection, cluster_attack_counts, container_unique_attack_counts, container_attack_counts, attack_type_counts, filters = process_data(df)
        # url = st.sidebar.multiselect(
        #     "Select URL:",
        #     options=df["URL"].unique(),
        #     # default=df["URL"].unique(),
        # )
        # Get unique values and add "Select All" option
        cluster_attack_counts_chart = plot_chart_cluster_most_unique_attack(cluster_attack_counts)
        container_unique_attack_counts_chart = plot_chart_container_most_unique_attack(container_unique_attack_counts)
        container_attack_counts_chart = plot_chart_container_most_attacks(container_attack_counts)
        attack_type_counts_chart = plot_top_attack_types(attack_type_counts)
        left_col, right_col = st.columns(2)
        with left_col:
            if cluster_attack_counts_chart is not None:
                st.plotly_chart(cluster_attack_counts_chart)
        with right_col:
            if container_unique_attack_counts_chart is not None:
                st.plotly_chart(container_unique_attack_counts_chart)
        
        # Display data summary
        df_summary = create_summary_dataframe(df_selection)
        st.dataframe(df_summary)
        data_summary = generate_message_content(df_summary)
        
        left_col, right_col = st.columns(2)
        with left_col:
            if container_attack_counts_chart is not None:
                st.plotly_chart(container_attack_counts_chart)
        with right_col:
            if attack_type_counts_chart is not None:
                st.plotly_chart(attack_type_counts_chart)
        st.subheader("Runtime Events Details")
        
        st.dataframe(df_selection, use_container_width=True)
        
        if st.button("Export to PDF"):
            with st.spinner("Generating PDF..."):
                fig_list = [cluster_attack_counts_chart, container_unique_attack_counts_chart, container_attack_counts_chart, attack_type_counts_chart]
                pdf_data = export_pdf.generate_pdf(fig_list, filters, "Runtime")
                if pdf_data:
                    st.download_button(
                        label="Download PDF",
                        data=pdf_data,
                        file_name=f"Runtime Events Report - {datetime.now().strftime('%A, %d %B %Y')}",
                        mime="application/pdf",
                    )
                else:
                    st.error("Failed to generate PDF. Please check the logs.")
        
        if st.button("Generate Cluster Summary Report"):
            with st.spinner("Generating Excel Cluster..."):
                filename = generate_excel_cluster(df)
                if filename is not None:
                    with open(filename, "rb") as f:
                        file_data = f.read()
                    st.download_button("Download Cluster Summary Report", data=file_data, file_name=f"Cluster Report Summary - {datetime.now().strftime('%A, %d %B %Y')}", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.success("Excel Cluster exported successfully.")
                else:
                    st.error("Failed to generate Excel Cluster. Please check the logs.")
    else:   
        st.warning("File does not exist.")
else:
    st.warning("No filename provided. Please go back to the Home Page to generate data.")